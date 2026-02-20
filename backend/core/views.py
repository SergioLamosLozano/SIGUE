import zipfile
import io
from rest_framework_simplejwt.views import TokenObtainPairView
from .serializers import CustomTokenObtainPairSerializer
from rest_framework import generics
from rest_framework.permissions import AllowAny, IsAdminUser
from rest_framework import permissions
from rest_framework import viewsets
from .serializers import RegisterSerializer, UserSerializer, UserAdminSerializer
from .serializers import RegisterSerializer, UserSerializer, UserAdminSerializer, LugarEventoSerializer, ProgramaSerializer
from .models import CustomUser, LugarEvento, Programa, Evento, EstudianteActivoUnivalle

class CustomTokenObtainPairView(TokenObtainPairView):
    """
    Vista personalizada para obtener el par de tokens (Access + Refresh).
    Usa CustomTokenObtainPairSerializer para devolver info extra del usuario.
    """
    serializer_class = CustomTokenObtainPairSerializer

class RegisterView(generics.CreateAPIView):
    """
    Vista para registrar nuevos usuarios.
    Permite acceso a cualquier persona (AllowAny) para que puedan registrarse.
    """
    queryset = CustomUser.objects.all()
    # Permitir a cualquier usuario (incluso no autenticado) acceder a este endpoint
    permission_classes = (AllowAny,)
    serializer_class = RegisterSerializer

class UserUpdateView(generics.RetrieveUpdateAPIView):
    """
    Vista para que el usuario autenticado lea y actualice su propio perfil.
    Solo permite acceso a usuarios autenticados.
    """
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_object(self):
        """
        Retorna el objeto usuario corresondiente al usuario que hace la petición (request.user).
        Esto asegura que un usuario solo pueda editar su propio perfil.
        """
        return self.request.user

class UserViewSet(viewsets.ModelViewSet):
    """
    ViewSet para la gestión completa (CRUD) de usuarios por parte de los administradores.
    Solo accesible por usuarios autenticados que también son Administradores (IsAdminUser).
    """
    # Consulta optimizada ordenando por nombre
    queryset = CustomUser.objects.all().order_by('full_name')
    serializer_class = UserAdminSerializer
    # Requiere autenticación y rol de administrador
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

class VerifyEmailView(APIView):
    """
    Endpoint para verificar el correo electrónico mediante código de 4 dígitos.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        id_usuario = request.data.get('id')
        code = request.data.get('code')

        if not id_usuario or not code:
            return Response({'error': 'ID y Código son obligatorios'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = CustomUser.objects.get(id=id_usuario)
        except CustomUser.DoesNotExist:
             return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        if user.is_active:
             return Response({'message': 'El usuario ya está activo'}, status=status.HTTP_200_OK)

        if user.verification_code == code:
            user.is_active = True
            user.verification_code = None # Limpiar código
            user.save()
            return Response({'message': 'Cuenta verificada exitosamente'}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Código incorrecto'}, status=status.HTTP_400_BAD_REQUEST)
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from .models import Asistente, CodigoQR, Evento, Inscripcion, Programa, EstudianteActivoUnivalle, GeneratedCertificate
from .serializers import (
    AsistenteSerializer, CodigoQRSerializer, EventoSerializer, 
    InscripcionSerializer, ProgramaSerializer, EstudianteActivoSerializer
)
import pandas as pd
from django.utils import timezone
from .email_utils import enviar_codigos_qr_email
import uuid
from django.core.files.base import ContentFile
import qrcode
from io import BytesIO
from django.db.models import Q
from django.core.exceptions import ValidationError
from django.conf import settings
from django.core.mail import send_mail
import json # Import json for robust handling

# -----------------------------------------------------------------------------
# EVENTO VIEWSET
# -----------------------------------------------------------------------------

class EventoViewSet(viewsets.ModelViewSet):
    """
    Controlador principal para la gestión de Eventos.
    Permite CRUD de eventos y acciones extra como inscripciones, generación de QRs y reportes.
    """
    queryset = Evento.objects.all().order_by('-fecha')
    serializer_class = EventoSerializer
    # Requiere autenticación para cualquier operación
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """
        Filtra los eventos según el rol del usuario:
        - Administrador: Ve todos los eventos.
        - Docente: Ve eventos aprobados + sus propios eventos (pendientes o aprobados).
        - Estudiante/Otros: Solo ven eventos aprobados.
        """
        user = self.request.user
        if not user.is_authenticated:
            return Evento.objects.none()

        if user.role == 'Administrador':
            return Evento.objects.all().order_by('-fecha')
        
        # Base query: Eventos aprobados
        queryset = Evento.objects.filter(estado='APROBADO')

        if user.role in ['Coordinador', 'Docente']:
            # Coordinadores y Docentes ven eventos aprobados + los suyos propios (pendientes/rechazados)
            mis_eventos = Evento.objects.filter(creado_por=user)
            queryset = queryset | mis_eventos
        
        # Estudiantes solo ven eventos aprobados (comportamiento default)
        
        return queryset.distinct().order_by('-fecha')
    
    # Imports locales para exportación CSV
    import csv
    from django.http import HttpResponse

    def create(self, request, *args, **kwargs):
        """Override create to include emails_enviados in the response."""
        self._emails_enviados = 0
        response = super().create(request, *args, **kwargs)
        if self._emails_enviados > 0:
            response.data['emails_enviados'] = self._emails_enviados
        return response

    def perform_create(self, serializer):
        """
        Asigna el creador.
        - Administrador: Crea eventos APROBADOS.
        - Docente/Coordinador: Crea eventos PENDIENTES (requieren aprobación del admin).
        - Estudiante: No puede crear eventos.
        
        Si enviar_difusion=true, envía correos a estudiantes de los programas seleccionados.
        """
        user = self.request.user
        
        from rest_framework.exceptions import PermissionDenied
        if user.role == 'Estudiante':
            raise PermissionDenied("No tienes permisos para crear eventos.")

        estado = 'APROBADO' if user.role == 'Administrador' else 'PENDIENTE'
        evento = serializer.save(creado_por=user, estado=estado)
        
        # Enviar difusión automática si se solicitó
        enviar_difusion = self.request.data.get('enviar_difusion', 'false')
        if enviar_difusion == 'true' and evento.programas_dirigidos.exists():
            self._emails_enviados = self._enviar_difusion_evento(evento)

    def _enviar_difusion_evento(self, evento):
        """
        Envía correos de difusión a todos los estudiantes activos 
        de los programas asociados al evento con un diseño HTML profesional.
        Retorna el número de correos enviados exitosamente.
        """
        from django.core.mail import EmailMultiAlternatives
        from django.conf import settings
        from email.mime.image import MIMEImage
        import base64
        
        programa_ids = evento.programas_dirigidos.values_list('id', flat=True)
        estudiantes = EstudianteActivoUnivalle.objects.filter(programa_id__in=programa_ids)
        
        emails_enviados = 0
        
        # Preparar el flyer si existe
        has_flyer = evento.flyer_data is not None and len(evento.flyer_data) > 0
        flyer_cid = 'flyer_evento'  # Content-ID para embeber en HTML
        
        for estudiante in estudiantes:
            if estudiante.correo:
                try:
                    fecha_str = evento.fecha.strftime('%d/%m/%Y a las %H:%M')
                    
                    # Texto plano como fallback
                    mensaje_texto = f"""
Hola {estudiante.nombre},

Te invitamos a participar en: {evento.titulo}

📅 Fecha: {fecha_str}
📍 Lugar: {evento.lugar}

{evento.descripcion}

Inscríbete en el sistema SIGUE para participar.

Saludos,
Universidad del Valle - SIGUE
"""
                    
                    # Sección del flyer en HTML (condicionalmente mostrar imagen o placeholder)
                    if has_flyer:
                        flyer_html = f'''
        <tr>
            <td style="padding: 0 30px 25px; text-align: center;">
                <img src="cid:{flyer_cid}" alt="Flyer del evento" style="max-width: 100%; height: auto; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
            </td>
        </tr>
'''
                    else:
                        flyer_html = '''
        <tr>
            <td style="padding: 0 30px 25px; text-align: center;">
                <div style="background: #f8f9fa; border: 2px dashed #ddd; border-radius: 12px; padding: 40px 20px; color: #888;">
                    <p style="margin: 0; font-size: 14px;">🖼️ <em>[Próximamente más información del evento]</em></p>
                </div>
            </td>
        </tr>
'''
                    
                    # Template HTML profesional con flyer embebido
                    mensaje_html = f'''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f4;">
    <table width="100%" cellpadding="0" cellspacing="0" style="max-width: 600px; margin: 0 auto; background-color: #ffffff;">
        
        <!-- Header con colores institucionales Universidad del Valle -->
        <tr>
            <td style="background: linear-gradient(135deg, #1a5f2a 0%, #2d8a3e 50%, #c41e3a 100%); padding: 30px 20px; text-align: center;">
                <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 700; text-shadow: 1px 1px 2px rgba(0,0,0,0.2);">
                    🎓 {evento.titulo}
                </h1>
                <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0; font-size: 14px;">
                    Universidad del Valle - Sistema SIGUE
                </p>
            </td>
        </tr>
        
        <!-- Saludo personalizado -->
        <tr>
            <td style="padding: 30px 30px 20px;">
                <p style="color: #333; font-size: 16px; margin: 0; line-height: 1.6;">
                    Hola <strong>{estudiante.nombre}</strong>,
                </p>
                <p style="color: #555; font-size: 15px; margin: 15px 0 0; line-height: 1.6;">
                    ¡Te invitamos a ser parte de una experiencia única! No te pierdas este evento especial que hemos preparado para ti y toda la comunidad universitaria.
                </p>
            </td>
        </tr>
        
        <!-- Detalles del evento -->
        <tr>
            <td style="padding: 0 30px;">
                <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(to right, #f8f9fa, #e9ecef); border-radius: 12px; border-left: 4px solid #1a5f2a;">
                    <tr>
                        <td style="padding: 25px;">
                            <table width="100%" cellpadding="0" cellspacing="0">
                                <tr>
                                    <td style="padding: 8px 0;">
                                        <span style="font-size: 20px;">📅</span>
                                        <span style="color: #333; font-size: 15px; margin-left: 10px;">
                                            <strong>Fecha:</strong> {fecha_str}
                                        </span>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 8px 0;">
                                        <span style="font-size: 20px;">📍</span>
                                        <span style="color: #333; font-size: 15px; margin-left: 10px;">
                                            <strong>Lugar:</strong> {evento.lugar}
                                        </span>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
        
        <!-- Descripción del evento -->
        <tr>
            <td style="padding: 25px 30px;">
                <p style="color: #444; font-size: 14px; line-height: 1.7; margin: 0; text-align: justify;">
                    {evento.descripcion}
                </p>
            </td>
        </tr>
        
        <!-- Flyer del evento (embebido o placeholder) -->
        {flyer_html}
        
        <!-- Botón de inscripción -->
        <tr>
            <td style="padding: 0 30px 30px; text-align: center;">
                <a href="#" style="display: inline-block; background: linear-gradient(135deg, #1a5f2a 0%, #2d8a3e 100%); color: #ffffff; text-decoration: none; padding: 16px 40px; border-radius: 50px; font-size: 16px; font-weight: 600; box-shadow: 0 4px 15px rgba(26, 95, 42, 0.3); transition: all 0.3s;">
                    ✨ Inscríbete en SIGUE
                </a>
                <p style="color: #888; font-size: 12px; margin: 15px 0 0;">
                    Haz clic para registrarte y asegurar tu participación
                </p>
            </td>
        </tr>
        
        <!-- Línea separadora -->
        <tr>
            <td style="padding: 0 30px;">
                <hr style="border: none; border-top: 1px solid #eee; margin: 0;">
            </td>
        </tr>
        
        <!-- Footer -->
        <tr>
            <td style="padding: 25px 30px; text-align: center; background: #f8f9fa;">
                <p style="color: #666; font-size: 14px; margin: 0 0 10px;">
                    ¡Te esperamos! 🎉
                </p>
                <p style="color: #1a5f2a; font-size: 16px; font-weight: 600; margin: 0;">
                    Universidad del Valle - SIGUE
                </p>
                <p style="color: #999; font-size: 11px; margin: 15px 0 0;">
                    Sistema Integrado de Gestión de Eventos
                </p>
            </td>
        </tr>
        
    </table>
</body>
</html>
'''
                    
                    # Crear email con alternativas (texto plano + HTML)
                    email = EmailMultiAlternatives(
                        subject=f'📣 Invitación: {evento.titulo}',
                        body=mensaje_texto.strip(),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[estudiante.correo]
                    )
                    email.attach_alternative(mensaje_html, "text/html")
                    email.mixed_subtype = 'related'  # Necesario para embeber imágenes inline
                    
                    # Adjuntar el flyer embebido (inline) si existe
                    if has_flyer:
                        # Crear imagen embebida con Content-ID
                        mime_image = MIMEImage(bytes(evento.flyer_data))
                        mime_image.add_header('Content-ID', f'<{flyer_cid}>')
                        mime_image.add_header('Content-Disposition', 'inline', filename=evento.flyer_filename or 'flyer.png')
                        email.attach(mime_image)
                        
                        # También adjuntar como archivo descargable
                        email.attach(
                            evento.flyer_filename or 'flyer.png',
                            bytes(evento.flyer_data),
                            evento.flyer_content_type or 'image/png'
                        )
                    
                    email.send(fail_silently=False)
                    emails_enviados += 1
                    
                except Exception as e:
                    # Log error but continue
                    print(f"Error enviando email a {estudiante.correo}: {e}")
        
        return emails_enviados

    @action(detail=True, methods=['post'])
    def aprobar(self, request, pk=None):
        """Permite a un administrador aprobar un evento pendiente."""
        if request.user.role != 'Administrador':
            return Response({'error': 'No tienes permisos para realizar esta acción'}, status=status.HTTP_403_FORBIDDEN)
        
        evento = self.get_object()
        evento.estado = 'APROBADO'
        evento.save()
        return Response({'message': 'Evento aprobado exitosamente'})

    @action(detail=True, methods=['post'])
    def unirse(self, request, pk=None):
        """
        Permite a un usuario inscribirse a un evento específico.
        Verifica que no esté ya inscrito.
        """
        evento = self.get_object()
        usuario = request.user
        
        if Inscripcion.objects.filter(evento=evento, usuario=usuario).exists():
            return Response({'message': 'Ya estás inscrito en este evento.'}, status=status.HTTP_400_BAD_REQUEST)
        
        Inscripcion.objects.create(evento=evento, usuario=usuario)
        return Response({'message': 'Inscripción exitosa.'}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'])
    def mis_eventos(self, request):
        """Devuelve la lista de eventos en los que el usuario actual está inscrito."""
        inscripciones = Inscripcion.objects.filter(usuario=request.user)
        eventos = [ins.evento for ins in inscripciones]
        serializer = EventoSerializer(eventos, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def inscritos(self, request, pk=None):
        """Devuelve la lista de personas inscritas a un evento específico."""
        evento = self.get_object()
        inscripciones = evento.inscripciones.all().select_related('usuario')
        serializer = InscripcionSerializer(inscripciones, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def estadisticas(self, request, pk=None):
        """
        Calcula estadísticas del evento:
        - Total inscritos
        - Asistencia real (basada en QRs de entrada usados)
        - Refrigerios entregados
        - Desglose por dependencia
        """
        evento = self.get_object()
        total_inscritos = evento.inscripciones.count()
        
        # Estadísticas de QRs
        qrs = CodigoQR.objects.filter(evento=evento)
        qrs_entregados = qrs.count()
        qrs_usados = qrs.filter(usado=True).count()
        
        # Asistencia Real = Códigos de tipo 'ENTRADA' que han sido usados
        asistentes_reales = qrs.filter(tipo_comida='ENTRADA', usado=True).count()
        
        # Estadísticas de Entregables (Desglose por Tipo)
        # Obtenemos todos los QRs que no son de entrada
        qrs_entregables = qrs.exclude(tipo_comida='ENTRADA')
        
        entregables_stats = {}
        
        # Iterar sobre los tipos de entregables encontrados
        # Usamos values('tipo_comida') para agrupar, pero como es SQLite/MySQL simple, 
        # podemos iterar o hacer consultas agregadas. 
        # Dado que no esperamos millones de registros, un loop simple sobre los tipos distintos funciona.
        tipos_distintos = qrs_entregables.values_list('tipo_comida', flat=True).distinct()
        
        total_generados_global = 0
        total_entregados_global = 0
        
        for tipo in tipos_distintos:
            # Filtrar QRs de este tipo
            qrs_tipo = qrs_entregables.filter(tipo_comida=tipo)
            generados = qrs_tipo.count()
            entregados = qrs_tipo.filter(usado=True).count()
            disponibles = generados - entregados
            
            entregables_stats[tipo] = {
                'generados': generados,
                'entregados': entregados,
                'disponibles': disponibles
            }
            
            total_generados_global += generados
            total_entregados_global += entregados

        # Desglose por Dependencia (solo de los que asistieron - Entrada)
        asistencia_qrs = qrs.filter(tipo_comida='ENTRADA', usado=True).select_related('usuario', 'asistente')
        
        dependencias_asistencia = {}
        for qr in asistencia_qrs:
            dep = "Sin Definir"
            if qr.usuario and qr.usuario.dependency:
                dep = qr.usuario.dependency
            elif qr.asistente and qr.asistente.sede:
                dep = qr.asistente.sede
            dep = dep.strip().title() if dep else "Sin Definir"
            dependencias_asistencia[dep] = dependencias_asistencia.get(dep, 0) + 1

        # Desglose por Dependencia (Inscritos)
        inscripciones = evento.inscripciones.all().select_related('usuario')
        dependencias_inscritos = {}
        
        for inscripcion in inscripciones:
            dep = "Sin Definir"
            if inscripcion.usuario and inscripcion.usuario.dependency:
                dep = inscripcion.usuario.dependency
            # Si hubiera asistentes legacy inscritos sin usuario (caso raro en modelo actual), manejar aquí
            
            dep = dep.strip().title() if dep else "Sin Definir"
            dependencias_inscritos[dep] = dependencias_inscritos.get(dep, 0) + 1

        return Response({
            'total_inscritos': total_inscritos,
            'asistentes_reales': asistentes_reales,
            'porcentaje_asistencia': (asistentes_reales / total_inscritos * 100) if total_inscritos > 0 else 0,
            
            # Globales
            'entregables_generados_total': total_generados_global,
            'entregables_entregados_total': total_entregados_global,
            'entregables_disponibles_total': total_generados_global - total_entregados_global,
            
            # Detalle
            'entregables_detalle': entregables_stats,
            
            'asistencia_por_dependencia': dependencias_asistencia, # Legacy key kept for compatibility if needed
            'inscritos_por_dependencia': dependencias_inscritos,
            'dependencias_comparativa': {
                'inscritos': dependencias_inscritos,
                'asistencia': dependencias_asistencia
            }
        })

    @action(detail=True, methods=['post'])
    def generar_qrs_masivo(self, request, pk=None):
        """
        Genera códigos QR para todos los inscritos en el evento.
        Crea QRs de Entrada y de los tipos de comida configurados.
        """
        evento = self.get_object()
        inscripciones = evento.inscripciones.all()
        generated_count = 0
        
        types = ['ENTRADA']
        
        try:
            # Verificar configuración de entregables
            detalles = evento.detalles_entregables
            
            # Robust JSON parsing: if it's a string, try to parse it
            if isinstance(detalles, str):
                try:
                    detalles = json.loads(detalles)
                except json.JSONDecodeError:
                    print(f"Error parsing detalles_entregables JSON for event {evento.id}: {detalles}")
                    detalles = {}
            
            detalles = detalles or {}
            items = detalles.get('items', [])
            
            # Si hay items personalizados, usarlos
            if isinstance(items, list) and len(items) > 0:
                types.extend([item for item in items if isinstance(item, str) and item.strip()])
                
            # Si no, usar lógica simple por defecto
            elif evento.requiere_entregable:
                types.append('ENTREGABLE')
                
            for inscripcion in inscripciones:
                user = inscripcion.usuario
                for tipo in types:
                    try:
                        _, created = CodigoQR.objects.get_or_create(
                            evento=evento,
                            usuario=user,
                            tipo_comida=tipo,
                            defaults={'asistente': None}
                        )
                        if created:
                            generated_count += 1
                    except Exception as e:
                        print(f"Error generating QR for user {user.id} type {tipo}: {e}")
            
            return Response({'message': f'Se generaron {generated_count} códigos QR nuevos.'})
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({'error': f"Error interno generando QRs: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            


    @action(detail=True, methods=['post'])
    def enviar_emails_evento(self, request, pk=None):
        """
        Envía los códigos QR por correo electrónico a todos los inscritos que tengan email.
        """
        evento = self.get_object()
        
        count = 0
        errors = 0
        error_details = []
        
        try:
             inscripciones = evento.inscripciones.all()
             for inscripcion in inscripciones:
                 user = inscripcion.usuario
                 
                 if not user.email:
                     continue

                 # Obtener QRs para este evento y usuario
                 qrs = CodigoQR.objects.filter(evento=evento, usuario=user)
                 
                 if qrs.exists():
                     # Clase adaptadora para que la función de envío de email funcione con el modelo User
                     # (Originalmente estaba hecha solo para Asistente legacy)
                     class AsistenteAdapter:
                         def __init__(self, u):
                             self.nombre_completo = u.full_name
                             self.correo = u.email
                             self.identificacion = u.id
                     
                     adapter = AsistenteAdapter(user)
                     
                     result = enviar_codigos_qr_email(adapter, evento, qrs)
                     if result is True:
                         count += 1
                     else:
                         errors += 1
                         error_details.append(f"{user.email}: {result}")
             
             return Response({
                 'message': f'Proceso finalizado. Emails enviados: {count}. Errores: {errors}',
                 'sent_count': count,
                 'error_count': errors,
                 'error_details': error_details
             })
        except Exception as e:
            print(f"CRITICAL ERROR in enviar_emails_evento: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def generar_certificados_masivo(self, request, pk=None):
        """
        Genera y envía certificados PDF a los asistentes que marcaron asistencia (asistio=True).
        Requiere que el evento tenga una plantilla PDF cargada.
        """
        evento = self.get_object()
        
        # 1. Actualizar plantilla si se envía una nueva en la petición
        plantilla = request.FILES.get('plantilla')
        if plantilla:
            evento.plantilla_certificado = plantilla
            evento.save()
            
        if not evento.plantilla_certificado:
            return Response({'error': 'No hay plantilla de certificado configurada para este evento.'}, status=status.HTTP_400_BAD_REQUEST)
        
        # 2. Filtrar solo los que asistieron
        inscripciones = evento.inscripciones.filter(asistio=True)
        
        generated_count = 0
        email_sent_count = 0
        errors = []
        
        from .utils import generar_certificado_pdf
        from django.core.mail import EmailMessage
        
        ruta_plantilla = evento.plantilla_certificado.path

        for inscripcion in inscripciones:
            user = inscripcion.usuario
            if not user.email:
                continue
                
            try:
                # Generar PDF en memoria usando la utilidad
                pdf_stream = generar_certificado_pdf(
                    user.full_name, 
                    user.id,
                    ruta_plantilla
                )
                
                if pdf_stream:
                    generated_count += 1
                    
                    # Preparar Email con adjunto
                    subject = f"Certificado de Asistencia - {evento.titulo}"
                    body = f"Hola {user.full_name},\n\nAdjunto encontrarás tu certificado de asistencia al evento '{evento.titulo}'.\n\n¡Gracias por participar!"
                    
                    from django.conf import settings
                    email = EmailMessage(
                        subject,
                        body,
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                    )
                    
                    # Adjuntar PDF
                    filename = f"Certificado_{user.full_name.replace(' ', '_')}.pdf"
                    email.attach(filename, pdf_stream.read(), 'application/pdf')
                    
                    # Enviar
                    email.send()
                    email_sent_count += 1
                    
            except Exception as e:
                errors.append(f"{user.email}: {str(e)}")
                
        return Response({
            'message': f'Proceso finalizado. Certificados generados: {generated_count}. Emails enviados: {email_sent_count}.',
            'errors': errors
        })

    @action(detail=True, methods=['post'])
    def ver_previsualizacion_certificado(self, request, pk=None):
        """
        Genera una vista previa del certificado con datos dummy para verificar alineación.
        """
        evento = self.get_object()
        from django.http import HttpResponse
        from .utils import generar_certificado_pdf
        
        # Guardar plantilla temporal si se envía
        plantilla_file = request.FILES.get('plantilla')
        if plantilla_file:
            evento.plantilla_certificado = plantilla_file
            evento.save()
            
        if not evento.plantilla_certificado:
            return Response({'error': 'No hay plantilla configurada.'}, status=status.HTTP_400_BAD_REQUEST)
        
        ruta_plantilla = evento.plantilla_certificado.path

        # Datos de prueba
        nombre_preview = "JUAN PEREZ (VISTA PREVIA)"
        doc_preview = "123456789"
        
        try:
            pdf_stream = generar_certificado_pdf(nombre_preview, doc_preview, ruta_plantilla)
            
            if pdf_stream:
                response = HttpResponse(pdf_stream.read(), content_type='application/pdf')
                response['Content-Disposition'] = 'inline; filename="certificado_preview.pdf"'
                return response
            else:
                return Response({'error': 'Error al generar PDF (stream vacío)'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def exportar_asistentes_excel(self, request, pk=None):
        """
        Genera un archivo .xlsx descargable con la lista de inscritos al evento.
        Usa openpyxl para compatibilidad con caracteres latinos.
        """
        import openpyxl
        from django.http import HttpResponse

        evento = self.get_object()

        # Filtrar solo inscripciones con asistencia CONFIRMADA
        inscripciones = evento.inscripciones.filter(asistio=True).select_related('usuario')

        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Asistentes"

        # Encabezados
        headers = ['Identificación', 'Nombre Completo', 'Email', 'Rol', 'Dependencia/Sede', 'Estado Asistencia', 'Fecha Inscripción']
        ws.append(headers)

        # Estilizar encabezados (negrita)
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Llenar filas
        for ins in inscripciones:
            user = ins.usuario
            estado = "Asistió" if ins.asistio else "Pendiente"
            fecha = ins.fecha_inscripcion.replace(tzinfo=None) if ins.fecha_inscripcion else ''

            ws.append([
                user.id,
                user.full_name,
                user.email or 'N/A',
                user.role,
                user.dependency or 'N/A',
                estado,
                fecha,
            ])

        # Autoajustar ancho de columnas
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="Asistentes_Evento_{evento.id}.xlsx"'

        # Guardar workbook directamente en la respuesta
        wb.save(response)
        return response

# -----------------------------------------------------------------------------
# ASISTENTE LEGACY VIEWSET
# -----------------------------------------------------------------------------

class AsistenteViewSet(viewsets.ModelViewSet):
    """
    CRUD para asistentes externos (Legacy).
    Permite importar masivamente desde Excel.
    """
    queryset = Asistente.objects.all()
    serializer_class = AsistenteSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['post'])
    def importar_excel(self, request):
        """
        Importa asistentes desde un archivo Excel (.xlsx).
        Crea automáticamente un QR de Entrada para los nuevos asistentes.
        """
        try:
            file = request.FILES.get('file')
            if not file:
                return Response({'error': 'No se proporcionó ningún archivo'}, status=status.HTTP_400_BAD_REQUEST)

            df = pd.read_excel(file)
            
            # Verificar columnas requeridas en el Excel
            required_columns = ['Nombre completo', 'Identificacion']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response(
                    {'error': f'Faltan columnas requeridas: {", ".join(missing_columns)}'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            created_count = 0
            updated_count = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    data = {
                        'identificacion': str(row['Identificacion']),
                        'nombre_completo': row['Nombre completo'],
                        'correo': row.get('Correo', ''),
                        'telefono': str(row.get('telefono', '')),
                        'sede': row.get('Sede', '')
                    }
                    
                    # Limpiar datos nulos (NaN)
                    data = {k: v if pd.notna(v) else '' for k, v in data.items()}

                    asistente, created = Asistente.objects.update_or_create(
                        identificacion=data['identificacion'],
                        defaults=data
                    )

                    if created:
                        created_count += 1
                        # Crear código QR de entrada
                        CodigoQR.objects.create(
                            asistente=asistente,
                            tipo_comida='ENTRADA'
                        )
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append(f"Fila {index + 2}: {str(e)}")

            return Response({
                'message': 'Proceso completado',
                'created': created_count,
                'updated': updated_count,
                'errors': errors
            })

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def generar_qr(self, request, pk=None):
        """Genera un QR manual para un asistente."""
        asistente = self.get_object()
        tipo_comida = request.data.get('tipo_comida')
        
        if not tipo_comida:
            return Response({'error': 'Tipo de comida requerido'}, status=status.HTTP_400_BAD_REQUEST)

        qr_obj, created = CodigoQR.objects.get_or_create(
            asistente=asistente,
            tipo_comida=tipo_comida
        )

        return Response(CodigoQRSerializer(qr_obj).data)
    
    @action(detail=True, methods=['post'])
    def enviar_qr_correo(self, request, pk=None):
        """Reenvía los códigos QR por correo a un asistente específico."""
        asistente = self.get_object()
        
        qrs = CodigoQR.objects.filter(asistente=asistente)
        
        if not qrs.exists():
            return Response({'error': 'El asistente no tiene códigos QR generados'}, status=status.HTTP_404_NOT_FOUND)
            
        if not asistente.correo:
            return Response({'error': 'El asistente no tiene correo registrado'}, status=status.HTTP_400_BAD_REQUEST)
            
        evento = qrs.first().evento
        if not evento:
             return Response({'error': 'No se encontró un evento asociado a estos códigos QR'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            enviar_codigos_qr_email(asistente, evento, qrs)
            return Response({'message': f'Códigos QR enviados a {asistente.correo}'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# -----------------------------------------------------------------------------
# CODIGO QR VIEWSET
# -----------------------------------------------------------------------------

class CodigoQRViewSet(viewsets.ModelViewSet):
    """
    CRUD para códigos QR y endpoint principal de ESCANEO.
    """
    queryset = CodigoQR.objects.all()
    serializer_class = CodigoQRSerializer
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['post'])
    def escanear(self, request):
        """
        Endpoint crítico para validar códigos QR.
        Recibe un 'codigo' que puede ser un UUID o una Cédula (entrada manual).
        """
        codigo = request.data.get('codigo')
        if not codigo:
            return Response({'error': 'Código requerido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            qr_obj = None
            # 1. Intentar buscar por UUID (QR estándar del sistema)
            try:
                qr_obj = CodigoQR.objects.select_related('usuario', 'asistente', 'evento').get(codigo=codigo)
            except (ValidationError, ValueError, CodigoQR.DoesNotExist):
                # 2. Si falla (ej. entrada manual de cédula), buscar por ID de Usuario o Asistente
                # Priorizamos encontrar un QR disponible (no usado) de tipo ENTRADA
                qr_obj = CodigoQR.objects.filter(
                    (Q(usuario__id=codigo) | Q(asistente__identificacion=codigo))
                ).select_related('usuario', 'asistente', 'evento').order_by('usado', 'fecha_creacion').first()

            if not qr_obj:
                 return Response({'error': 'Código o Identificación no válida'}, status=status.HTTP_404_NOT_FOUND)
            
            # Construir información de respuesta normalizada
            attendant_info = {}
            if qr_obj.usuario:
                attendant_info = {
                    'nombre_completo': qr_obj.usuario.full_name,
                    'identificacion': qr_obj.usuario.id,
                    'sede': qr_obj.usuario.dependency or 'N/A',
                    'email': qr_obj.usuario.email
                }
            elif qr_obj.asistente:
                attendant_info = {
                    'nombre_completo': qr_obj.asistente.nombre_completo,
                    'identificacion': qr_obj.asistente.identificacion,
                    'sede': qr_obj.asistente.sede or 'N/A',
                    'email': qr_obj.asistente.correo
                }
            else:
                attendant_info = {
                    'nombre_completo': 'Desconocido',
                    'identificacion': 'N/A',
                    'sede': 'N/A'
                }

            # Validar si ya fue usado
            if qr_obj.usado:
                return Response({
                    'status': 'error',
                    'message': f'Este código ya fue usado el {qr_obj.fecha_uso.strftime("%d/%m/%Y %H:%M") if qr_obj.fecha_uso else "previamente"}',
                    'asistente': attendant_info,
                    'tipo_comida': qr_obj.tipo_comida,
                    'evento': qr_obj.evento.titulo if qr_obj.evento else None
                }, status=status.HTTP_400_BAD_REQUEST)

            # MARCAR COMO USADO (Redimir)
            qr_obj.marcar_como_usado()
            
            return Response({
                'status': 'success',
                'message': 'Código validado exitosamente',
                'asistente': attendant_info,
                'tipo_comida': qr_obj.tipo_comida,
                'fecha_uso': qr_obj.fecha_uso,
                'evento': qr_obj.evento.titulo if qr_obj.evento else None
            })

        except Exception as e:
            print(f"ERROR CRÍTICO en escanear: {str(e)}")
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# -----------------------------------------------------------------------------
# PROGRAMA VIEWSET (Read-only for listing programs)
# -----------------------------------------------------------------------------

# -----------------------------------------------------------------------------
# PROGRAMA & LOCATION VIEWSETS
# -----------------------------------------------------------------------------

class LugarEventoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de Lugares/Ubicaciones (CRUD).
    """
    queryset = LugarEvento.objects.all().order_by('descripcion')
    serializer_class = LugarEventoSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_permissions(self):
        """
        Solo administradores pueden crear, editar o eliminar lugares.
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdminUser()]
        return [permissions.IsAuthenticated()]


class ProgramaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión completa de Programas Académicos (CRUD).
    - Administradores: pueden crear, editar y eliminar programas.
    - Otros usuarios autenticados: solo pueden listar y ver programas.
    """
    queryset = Programa.objects.all().order_by('descripcion')
    serializer_class = ProgramaSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_permissions(self):
        """
        Solo administradores pueden crear, editar o eliminar programas.
        """
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [permissions.IsAuthenticated(), IsAdminUser()]
        return [permissions.IsAuthenticated()]


class EstudianteActivoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet para listar Estudiantes Activos con paginación.
    Solo lectura - los estudiantes se cargan vía Excel.
    """
    queryset = EstudianteActivoUnivalle.objects.all().select_related('programa').order_by('nombre')
    serializer_class = EstudianteActivoSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Filtro por programa si se especifica
        programa_id = self.request.query_params.get('programa', None)
        if programa_id:
            queryset = queryset.filter(programa_id=programa_id)
        # Búsqueda por nombre o código
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                models.Q(nombre__icontains=search) | 
                models.Q(codigo_estudiante__icontains=search)
            )
        return queryset


# -----------------------------------------------------------------------------
# EXCEL UPLOAD FOR ACTIVE STUDENTS
# -----------------------------------------------------------------------------

class CargarEstudiantesExcelView(APIView):
    """
    Endpoint para cargar estudiantes activos desde archivo Excel.
    Solo accesible por administradores.
    
    Columnas esperadas en el Excel:
    - Codigo (mapeado a codigo_estudiante)
    - Apellidos + Nombres (concatenados en nombre)
    - Email (mapeado a correo)
    - Programa Académico (mapeado a programa_id)
    """
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        if 'archivo' not in request.FILES:
            return Response(
                {'error': 'No se proporcionó archivo Excel'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        archivo = request.FILES['archivo']
        
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(archivo, read_only=True)
            ws = wb.active
            
            # Buscar la fila de encabezados (la que contiene "Código" o "Codigo")
            # El archivo Excel de la universidad tiene encabezados en filas superiores
            header_row = None
            headers = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
                row_str = ' '.join([str(cell).lower() if cell else '' for cell in row])
                if 'codigo' in row_str or 'código' in row_str:
                    header_row = row_idx
                    headers = [str(cell).strip().lower() if cell else '' for cell in row]
                    break
            
            if not header_row:
                return Response(
                    {'error': 'No se encontró la fila de encabezados con "Código". Asegúrate de que el Excel tenga una fila con columnas como: Código, Apellidos, Nombres, Email, Programa'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Mapeo de columnas esperadas
            col_map = {}
            for idx, header in enumerate(headers):
                if 'codigo' in header or 'código' in header:
                    col_map['codigo'] = idx
                elif 'apellido' in header:
                    col_map['apellidos'] = idx
                elif 'nombre' in header and 'programa' not in header:
                    col_map['nombres'] = idx
                elif 'email' in header or 'correo' in header:
                    col_map['email'] = idx
                elif 'programa' in header and 'académico' in header:
                    col_map['programa'] = idx
                elif 'programa' in header and 'acade' in header:
                    col_map['programa'] = idx
                elif 'programa' in header and 'nombre' not in header:
                    col_map['programa'] = idx
            
            if 'codigo' not in col_map:
                return Response(
                    {'error': 'No se encontró la columna "Código" en el Excel'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            estudiantes_creados = 0
            estudiantes_actualizados = 0
            errores = []
            
            # Leer datos a partir de la fila siguiente a los encabezados
            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                try:
                    codigo = str(row[col_map['codigo']]).strip() if row[col_map['codigo']] else None
                    if not codigo:
                        continue
                    
                    # Construir nombre completo
                    apellidos = str(row[col_map.get('apellidos', 0)] or '').strip()
                    nombres = str(row[col_map.get('nombres', 0)] or '').strip()
                    nombre_completo = f"{apellidos} {nombres}".strip() or codigo
                    
                    # Email
                    email = str(row[col_map.get('email', 0)] or '').strip() or None
                    
                    # Programa
                    programa_id = row[col_map.get('programa')] if 'programa' in col_map else None
                    programa = None
                    if programa_id:
                        try:
                            programa = Programa.objects.get(id=int(programa_id))
                        except (Programa.DoesNotExist, ValueError):
                            pass
                    
                    # Insert or Update estudiante
                    obj, created = EstudianteActivoUnivalle.objects.update_or_create(
                        codigo_estudiante=codigo,
                        defaults={
                            'nombre': nombre_completo,
                            'correo': email,
                            'programa': programa
                        }
                    )
                    
                    if created:
                        estudiantes_creados += 1
                    else:
                        estudiantes_actualizados += 1
                        
                except Exception as e:
                    errores.append(f"Fila {row_idx}: {str(e)}")
            
            wb.close()
            
            return Response({
                'message': 'Carga completada',
                'estudiantes_creados': estudiantes_creados,
                'estudiantes_actualizados': estudiantes_actualizados,
                'errores': errores[:10]  # Solo primeros 10 errores
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error procesando el archivo: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# -----------------------------------------------------------------------------
# EMAIL DIFFUSION FOR EVENTS - Add action to EventoViewSet
# -----------------------------------------------------------------------------

# NOTE: The enviar_difusion action should be added to EventoViewSet above
# For now, creating a standalone view that can be called with event ID

class EnviarDifusionEventoView(APIView):
    """
    Envía correos de difusión/promoción de un evento a todos los estudiantes
    activos que pertenecen a los programas seleccionados en 'A quién va dirigido'.
    Procesa de forma síncrona y retorna el conteo real de correos enviados.
    """
    permission_classes = [permissions.IsAuthenticated, IsAdminUser]
    
    def post(self, request, evento_id):
        from email.mime.image import MIMEImage
        from django.core.mail import EmailMultiAlternatives
        
        try:
            evento = Evento.objects.get(id=evento_id)
        except Evento.DoesNotExist:
            return Response({'error': 'Evento no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        programas = evento.programas_dirigidos.all()
        
        if not programas.exists():
            return Response(
                {'error': 'El evento no tiene programas dirigidos configurados'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener estudiantes de los programas seleccionados con email válido
        estudiantes = EstudianteActivoUnivalle.objects.filter(
            programa__in=programas,
            correo__isnull=False
        ).exclude(correo='')
        
        if not estudiantes.exists():
            return Response(
                {'error': 'No hay estudiantes con email en los programas seleccionados'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Preparar el flyer si existe
        has_flyer = evento.flyer_data is not None and len(evento.flyer_data) > 0
        flyer_cid = 'flyer_evento'
        
        asunto = f"📢 Invitación: {evento.titulo}"
        fecha_str = evento.fecha.strftime("%d/%m/%Y a las %H:%M")
        
        emails_enviados = 0
        errores = []
        
        for estudiante in estudiantes:
            try:
                # Texto plano como fallback
                mensaje_texto = f"""
Hola {estudiante.nombre},

Te invitamos a participar en: {evento.titulo}

📅 Fecha: {fecha_str}
📍 Lugar: {evento.lugar}

{evento.descripcion}

Inscríbete en el sistema SIGUE para participar.

Saludos,
Universidad del Valle - SIGUE
"""
                
                # Sección del flyer en HTML
                if has_flyer:
                    flyer_html = f'''
        <tr>
            <td style="padding: 0 30px 25px; text-align: center;">
                <img src="cid:{flyer_cid}" alt="Flyer del evento" style="max-width: 100%; height: auto; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
            </td>
        </tr>
'''
                else:
                    flyer_html = '''
        <tr>
            <td style="padding: 0 30px 25px; text-align: center;">
                <div style="background: #f8f9fa; border: 2px dashed #ddd; border-radius: 12px; padding: 40px 20px; color: #888;">
                    <p style="margin: 0; font-size: 14px;">🖼️ <em>[Próximamente más información del evento]</em></p>
                </div>
            </td>
        </tr>
'''
                
                # Template HTML profesional
                mensaje_html = f'''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f4;">
    <table width="100%" cellpadding="0" cellspacing="0" style="max-width: 600px; margin: 0 auto; background-color: #ffffff;">
        
        <!-- Header con colores institucionales Universidad del Valle -->
        <tr>
            <td style="background: linear-gradient(135deg, #1a5f2a 0%, #2d8a3e 50%, #c41e3a 100%); padding: 30px 20px; text-align: center;">
                <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 700; text-shadow: 1px 1px 2px rgba(0,0,0,0.2);">
                    🎓 {evento.titulo}
                </h1>
                <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0; font-size: 14px;">
                    Universidad del Valle - Sistema SIGUE
                </p>
            </td>
        </tr>
        
        <!-- Saludo personalizado -->
        <tr>
            <td style="padding: 30px 30px 20px;">
                <p style="color: #333; font-size: 16px; margin: 0; line-height: 1.6;">
                    Hola <strong>{estudiante.nombre}</strong>,
                </p>
                <p style="color: #555; font-size: 15px; margin: 15px 0 0; line-height: 1.6;">
                    ¡Te invitamos a ser parte de una experiencia única! No te pierdas este evento especial que hemos preparado para ti y toda la comunidad universitaria.
                </p>
            </td>
        </tr>
        
        <!-- Detalles del evento -->
        <tr>
            <td style="padding: 0 30px;">
                <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(to right, #f8f9fa, #e9ecef); border-radius: 12px; border-left: 4px solid #1a5f2a;">
                    <tr>
                        <td style="padding: 25px;">
                            <table width="100%" cellpadding="0" cellspacing="0">
                                <tr>
                                    <td style="padding: 8px 0;">
                                        <span style="font-size: 20px;">📅</span>
                                        <span style="color: #333; font-size: 15px; margin-left: 10px;">
                                            <strong>Fecha:</strong> {fecha_str}
                                        </span>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 8px 0;">
                                        <span style="font-size: 20px;">📍</span>
                                        <span style="color: #333; font-size: 15px; margin-left: 10px;">
                                            <strong>Lugar:</strong> {evento.lugar}
                                        </span>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
        
        <!-- Descripción del evento -->
        <tr>
            <td style="padding: 25px 30px;">
                <p style="color: #444; font-size: 14px; line-height: 1.7; margin: 0; text-align: justify;">
                    {evento.descripcion}
                </p>
            </td>
        </tr>
        
        <!-- Flyer del evento (dinámico) -->
        {flyer_html}
        
        <!-- Botón de inscripción -->
        <tr>
            <td style="padding: 0 30px 30px; text-align: center;">
                <a href="https://ejecafetero.univalle.edu.co/" style="display: inline-block; background: linear-gradient(135deg, #1a5f2a 0%, #2d8a3e 100%); color: #ffffff; text-decoration: none; padding: 16px 40px; border-radius: 50px; font-size: 16px; font-weight: 600; box-shadow: 0 4px 15px rgba(26, 95, 42, 0.3);">
                    ✨ Inscríbete en SIGUE
                </a>
                <p style="color: #888; font-size: 12px; margin: 15px 0 0;">
                    Haz clic para registrarte y asegurar tu participación
                </p>
            </td>
        </tr>
        
        <!-- Línea separadora -->
        <tr>
            <td style="padding: 0 30px;">
                <hr style="border: none; border-top: 1px solid #eee; margin: 0;">
            </td>
        </tr>
        
        <!-- Footer -->
        <tr>
            <td style="padding: 25px 30px; text-align: center; background: #f8f9fa;">
                <p style="color: #666; font-size: 14px; margin: 0 0 10px;">
                    ¡Te esperamos! 🎉
                </p>
                <p style="color: #1a5f2a; font-size: 16px; font-weight: 600; margin: 0;">
                    Universidad del Valle - SIGUE
                </p>
                <p style="color: #999; font-size: 11px; margin: 15px 0 0;">
                    Sistema Integrado de Gestión de Eventos
                </p>
            </td>
        </tr>
        
    </table>
</body>
</html>
'''
                
                # Crear email con alternativas (texto plano + HTML)
                email = EmailMultiAlternatives(
                    subject=asunto,
                    body=mensaje_texto.strip(),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[estudiante.correo]
                )
                email.attach_alternative(mensaje_html, "text/html")
                email.mixed_subtype = 'related'
                
                # Adjuntar el flyer embebido (inline) si existe
                if has_flyer:
                    mime_image = MIMEImage(bytes(evento.flyer_data))
                    mime_image.add_header('Content-ID', f'<{flyer_cid}>')
                    mime_image.add_header('Content-Disposition', 'inline', filename=evento.flyer_filename or 'flyer.png')
                    email.attach(mime_image)
                    
                    email.attach(
                        evento.flyer_filename or 'flyer.png',
                        bytes(evento.flyer_data),
                        evento.flyer_content_type or 'image/png'
                    )
                
                email.send(fail_silently=False)
                emails_enviados += 1
                
            except Exception as e:
                errores.append(f"{estudiante.correo}: {str(e)}")
        
        return Response({
            'message': 'Difusión completada',
            'emails_enviados': emails_enviados,
            'total_estudiantes': estudiantes.count(),
            'programas': [p.descripcion for p in programas],
            'errores': errores[:10]
        })


def _worker_send_difusion(evento_id, admin_email):
    """Worker que ejecuta el envío masivo de difusión en un hilo separado."""
    from email.mime.image import MIMEImage
    from django.core.mail import EmailMultiAlternatives
    from django.core.mail import send_mail as django_send_mail

    emails_enviados = 0
    errores = []
    evento_titulo = "Evento"

    try:
        evento = Evento.objects.get(id=evento_id)
        evento_titulo = evento.titulo
        programas = evento.programas_dirigidos.all()
        
        estudiantes = EstudianteActivoUnivalle.objects.filter(
            programa__in=programas,
            correo__isnull=False
        ).exclude(correo='')
        
        # Preparar el flyer si existe
        has_flyer = evento.flyer_data is not None and len(evento.flyer_data) > 0
        flyer_cid = 'flyer_evento'
        
        asunto = f"📢 Invitación: {evento.titulo}"
        fecha_str = evento.fecha.strftime("%d/%m/%Y a las %H:%M")
        
        for estudiante in estudiantes:
            try:
                # Texto plano como fallback
                mensaje_texto = f"""
Hola {estudiante.nombre},

Te invitamos a participar en: {evento.titulo}

📅 Fecha: {fecha_str}
📍 Lugar: {evento.lugar}

{evento.descripcion}

Inscríbete en el sistema SIGUE para participar.

Saludos,
Universidad del Valle - SIGUE
"""
                
                # Sección del flyer en HTML (condicionalmente mostrar imagen o placeholder)
                if has_flyer:
                    flyer_html = f'''
        <tr>
            <td style="padding: 0 30px 25px; text-align: center;">
                <img src="cid:{flyer_cid}" alt="Flyer del evento" style="max-width: 100%; height: auto; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1);">
            </td>
        </tr>
'''
                else:
                    flyer_html = '''
        <tr>
            <td style="padding: 0 30px 25px; text-align: center;">
                <div style="background: #f8f9fa; border: 2px dashed #ddd; border-radius: 12px; padding: 40px 20px; color: #888;">
                    <p style="margin: 0; font-size: 14px;">🖼️ <em>[Próximamente más información del evento]</em></p>
                </div>
            </td>
        </tr>
'''
                
                # Template HTML profesional
                mensaje_html = f'''
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f4f4;">
    <table width="100%" cellpadding="0" cellspacing="0" style="max-width: 600px; margin: 0 auto; background-color: #ffffff;">
        
        <!-- Header con colores institucionales Universidad del Valle -->
        <tr>
            <td style="background: linear-gradient(135deg, #1a5f2a 0%, #2d8a3e 50%, #c41e3a 100%); padding: 30px 20px; text-align: center;">
                <h1 style="color: #ffffff; margin: 0; font-size: 28px; font-weight: 700; text-shadow: 1px 1px 2px rgba(0,0,0,0.2);">
                    🎓 {evento.titulo}
                </h1>
                <p style="color: rgba(255,255,255,0.9); margin: 10px 0 0; font-size: 14px;">
                    Universidad del Valle - Sistema SIGUE
                </p>
            </td>
        </tr>
        
        <!-- Saludo personalizado -->
        <tr>
            <td style="padding: 30px 30px 20px;">
                <p style="color: #333; font-size: 16px; margin: 0; line-height: 1.6;">
                    Hola <strong>{estudiante.nombre}</strong>,
                </p>
                <p style="color: #555; font-size: 15px; margin: 15px 0 0; line-height: 1.6;">
                    ¡Te invitamos a ser parte de una experiencia única! No te pierdas este evento especial que hemos preparado para ti y toda la comunidad universitaria.
                </p>
            </td>
        </tr>
        
        <!-- Detalles del evento -->
        <tr>
            <td style="padding: 0 30px;">
                <table width="100%" cellpadding="0" cellspacing="0" style="background: linear-gradient(to right, #f8f9fa, #e9ecef); border-radius: 12px; border-left: 4px solid #1a5f2a;">
                    <tr>
                        <td style="padding: 25px;">
                            <table width="100%" cellpadding="0" cellspacing="0">
                                <tr>
                                    <td style="padding: 8px 0;">
                                        <span style="font-size: 20px;">📅</span>
                                        <span style="color: #333; font-size: 15px; margin-left: 10px;">
                                            <strong>Fecha:</strong> {fecha_str}
                                        </span>
                                    </td>
                                </tr>
                                <tr>
                                    <td style="padding: 8px 0;">
                                        <span style="font-size: 20px;">📍</span>
                                        <span style="color: #333; font-size: 15px; margin-left: 10px;">
                                            <strong>Lugar:</strong> {evento.lugar}
                                        </span>
                                    </td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                </table>
            </td>
        </tr>
        
        <!-- Descripción del evento -->
        <tr>
            <td style="padding: 25px 30px;">
                <p style="color: #444; font-size: 14px; line-height: 1.7; margin: 0; text-align: justify;">
                    {evento.descripcion}
                </p>
            </td>
        </tr>
        
        <!-- Flyer del evento (dinámico) -->
        {flyer_html}
        
        <!-- Botón de inscripción -->
        <tr>
            <td style="padding: 0 30px 30px; text-align: center;">
                <a href="https://ejecafetero.univalle.edu.co/" style="display: inline-block; background: linear-gradient(135deg, #1a5f2a 0%, #2d8a3e 100%); color: #ffffff; text-decoration: none; padding: 16px 40px; border-radius: 50px; font-size: 16px; font-weight: 600; box-shadow: 0 4px 15px rgba(26, 95, 42, 0.3);">
                    ✨ Inscríbete en SIGUE
                </a>
                <p style="color: #888; font-size: 12px; margin: 15px 0 0;">
                    Haz clic para registrarte y asegurar tu participación
                </p>
            </td>
        </tr>
        
        <!-- Línea separadora -->
        <tr>
            <td style="padding: 0 30px;">
                <hr style="border: none; border-top: 1px solid #eee; margin: 0;">
            </td>
        </tr>
        
        <!-- Footer -->
        <tr>
            <td style="padding: 25px 30px; text-align: center; background: #f8f9fa;">
                <p style="color: #666; font-size: 14px; margin: 0 0 10px;">
                    ¡Te esperamos! 🎉
                </p>
                <p style="color: #1a5f2a; font-size: 16px; font-weight: 600; margin: 0;">
                    Universidad del Valle - SIGUE
                </p>
                <p style="color: #999; font-size: 11px; margin: 15px 0 0;">
                    Sistema Integrado de Gestión de Eventos
                </p>
            </td>
        </tr>
        
    </table>
</body>
</html>
'''
                
                # Crear email con alternativas (texto plano + HTML)
                from django.core.mail import EmailMultiAlternatives
                email = EmailMultiAlternatives(
                    subject=asunto,
                    body=mensaje_texto.strip(),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[estudiante.correo]
                )
                email.attach_alternative(mensaje_html, "text/html")
                email.mixed_subtype = 'related'  # Necesario para embeber imágenes inline
                
                # Adjuntar el flyer embebido (inline) si existe
                if has_flyer:
                    # Crear imagen embebida con Content-ID
                    mime_image = MIMEImage(bytes(evento.flyer_data))
                    mime_image.add_header('Content-ID', f'<{flyer_cid}>')
                    mime_image.add_header('Content-Disposition', 'inline', filename=evento.flyer_filename or 'flyer.png')
                    email.attach(mime_image)
                    
                    # También adjuntar como archivo descargable
                    email.attach(
                        evento.flyer_filename or 'flyer.png',
                        bytes(evento.flyer_data),
                        evento.flyer_content_type or 'image/png'
                    )
                
                email.send(fail_silently=False)
                emails_enviados += 1
                
            except Exception as e:
                errores.append(f"{estudiante.correo}: {str(e)}")

    except Exception as e:
        print(f"Error crítico en hilo de difusión: {e}")
        errores.append(f"Error crítico: {str(e)}")

    # Enviar reporte al admin
    if admin_email:
        try:
            errores_str = "\n".join(errores[:10]) if errores else "Ninguno"
            django_send_mail(
                f'✅ Reporte de Difusión - {evento_titulo}',
                f'El envío masivo de difusión ha finalizado.\n\n'
                f'📊 Resumen:\n'
                f'  • Correos enviados: {emails_enviados}\n'
                f'  • Errores: {len(errores)}\n\n'
                f'Errores:\n{errores_str}\n\n'
                f'— Sistema SIGUE',
                settings.DEFAULT_FROM_EMAIL,
                [admin_email],
                fail_silently=True
            )
        except Exception as e:
            print(f"Error enviando reporte al admin: {e}")

    print(f"[HILO DIFUSIÓN] Finalizado — Enviados: {emails_enviados}, Errores: {len(errores)}")

# -----------------------------------------------------------------------------
# CERTIFICADOS VIEW
# -----------------------------------------------------------------------------

class UploadCertificateTemplateView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        import json
        print("Datos recibidos:", request.data) # Debug
        
        event_id = request.data.get('event')
        file_obj = request.data.get('image')
        # Frontend envía 'config_data' como string JSON
        config_data_str = request.data.get('config_data')

        if not event_id or not file_obj:
            return Response({"error": "Faltan datos (evento o imagen)"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # LÓGICA DE GUARDADO:
            # 1. Buscar el evento
            evento = Evento.objects.get(id=event_id)
            
            # 2. Actualizar la plantilla en el evento
            evento.plantilla_certificado = file_obj
            
            # 3. Guardar la configuración JSON si existe
            if config_data_str:
                try:
                    evento.config_certificado = json.loads(config_data_str)
                except json.JSONDecodeError:
                    print("Error decodificando JSON de configuración")
                    # No fallar del todo, pero loguear
            
            evento.save()
            
            return Response({"message": "Plantilla y configuración guardadas exitosamente", "id": evento.id}, status=status.HTTP_201_CREATED)
            
        except Evento.DoesNotExist:
             return Response({"error": "Evento no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Error subiendo plantilla: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class GenerateBulkCertificatesView(APIView):
    """
    Genera certificados masivos para todos los asistentes (asistio=True) de un evento.
    Usa la plantilla (Imagen) base y las coordenadas JSON guardadas en el evento.
    Guarda los PDFs generados como BLOBs en la BD.
    """
    def post(self, request, *args, **kwargs):
        import io
        import json
        from reportlab.pdfgen import canvas
        from reportlab.lib.utils import ImageReader
        from django.core.files.base import ContentFile

        event_id = request.data.get('event_id')
        if not event_id:
            return Response({"error": "Falta event_id"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            evento = Evento.objects.get(id=event_id)
            if not evento.plantilla_certificado:
                return Response({"error": "El evento no tiene plantilla configurada"}, status=status.HTTP_400_BAD_REQUEST)
            
            # BLINDAJE CONTRA CONFIGURACIÓN VACÍA
            if not evento.config_certificado:
                return Response({
                    "error": "La plantilla existe pero NO TIENE CONFIGURACIÓN (Coordenadas). "
                             "Por favor vuelve al Diseñador y dale 'Guardar' de nuevo para actualizarla."
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Obtener inscritos confirmados
            inscripciones = Inscripcion.objects.filter(evento=evento, asistio=True).select_related('usuario')
            if not inscripciones.exists():
                return Response({"message": "No hay asistentes confirmados para este evento."}, status=status.HTTP_200_OK)

            generated_count = 0
            
            # Obtener configuración (manejo robusto de JSON/Lista)
            raw_config = evento.config_certificado or []
            
            # Normalizar a lista de campos
            if isinstance(raw_config, str):
                try:
                    raw_config = json.loads(raw_config)
                except:
                    raw_config = []
            
            fields_config = []
            if isinstance(raw_config, list):
                fields_config = raw_config
            elif isinstance(raw_config, dict):
                # Si viene como objeto, buscamos 'fields' o intentamos convertir
                fields_config = raw_config.get('fields', [])
                if not fields_config and raw_config:
                    # Fallback por si acaso es formato dict antiguo
                    fields_config = [v for k, v in raw_config.items()]

            # Cargar la imagen plantilla en memoria una sola vez
            try:
                plantilla_path = evento.plantilla_certificado.path # FileSystem path
                bg_image = ImageReader(plantilla_path) 
                img_w, img_h = bg_image.getSize()
            except Exception as e:
                return Response({"error": f"Error leyendo plantilla de imagen: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            for ins in inscripciones:
                user = ins.usuario
                
                # Crear Buffer PDF
                buffer = io.BytesIO()
                c = canvas.Canvas(buffer, pagesize=(img_w, img_h))
                
                # 1. Dibujar Plantilla de Fondo
                c.drawImage(bg_image, 0, 0, width=img_w, height=img_h)
                
                # 2. Dibujar Textos según Configuración
                for field in fields_config:
                    try:
                        f_id = field.get('id', '')
                        f_type = field.get('type', 'text')
                        
                        # --- LÓGICA DE INYECCIÓN DE DATOS REALES ---
                        # 1. Definir texto por defecto (Placeholder o estático)
                        default_text = str(field.get('text', ''))
                        text_to_draw = default_text
                        
                        normalized_id = str(f_id).lower()
                        normalized_text = default_text.lower()

                        # 2. Reemplazo Dinámico (Prioridad Alta)
                        # Buscamos coincidencias tanto en el ID como en el contenido del texto placeholder
                        keys_nombre = ['nombre', 'name', 'student', 'estudiante']
                        keys_id = ['cedula', 'id', 'documento', 'cc', 'identificacion']

                        if any(x in normalized_id for x in keys_nombre) or any(k in normalized_text for k in keys_nombre):
                            text_to_draw = user.full_name.upper()
                        
                        elif any(x in normalized_id for x in keys_id) or any(k in normalized_text for k in keys_id):
                            text_to_draw = str(user.id)
                        
                        # Si no hay texto y no es imagen, saltar
                        if not text_to_draw and f_type == 'text':
                            continue

                        # --- Coordenadas y Estilos (CORRECCIÓN VISUAL) ---
                        # React envía porcentajes (0-100)
                        x_pct = float(field.get('x', 0))
                        y_pct = float(field.get('y', 0))
                        
                        if f_type == 'text':
                            font_size = int(field.get('fontSize', 12))
                            font_family = field.get('fontFamily', 'Helvetica')
                            
                            # Mapping fuentes
                            rl_font = 'Helvetica-Bold'
                            if 'Times' in font_family: rl_font = 'Times-Roman'
                            elif 'Courier' in font_family: rl_font = 'Courier'
                            
                            c.setFont(rl_font, font_size)
                            
                            # Conversión a Puntos (Points) con Corrección de Baseline
                            # X: Simple regla de tres
                            x_pos = img_w * (x_pct / 100.0)
                            
                            # Y: Invertido (0 abajo) - Ajuste PRO (FontSize * 1.15)
                            # Esto baja el texto para que no quede "flotando"
                            y_pos = img_h - (img_h * (y_pct / 100.0)) - (font_size * 1.15)

                            c.drawString(x_pos, y_pos, str(text_to_draw))
                            
                        # TODO: Lógica futura para imágenes (firmas) si f_type == 'image'
                    
                    except Exception as field_err:
                        print(f"Error pintando campo {field}: {field_err}")
                        continue
                
                c.showPage()
                c.save()
                
                # 3. Guardar en Base de Datos (BLOB)
                pdf_bytes = buffer.getvalue()
                filename = f"Certificado_{user.id}_{evento.id}.pdf"
                
                # Update or Create
                GeneratedCertificate.objects.update_or_create(
                    usuario=user,
                    evento=evento,
                    defaults={
                        'pdf_blob': pdf_bytes,
                        'filename': filename,
                        'content_type': 'application/pdf',
                        'created_at': timezone.now()
                    }
                )
                generated_count += 1
                buffer.close()

            # --- CORRECCIÓN CRÍTICA: RETORNO SEGURO ---
            return Response({
                "message": f"Proceso finalizado. {generated_count} certificados generados.",
                "generated_count": generated_count,
                "status": "success"
            }, status=status.HTTP_200_OK)

        except Evento.DoesNotExist:
            return Response({"error": "Evento no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"Error generando certificados: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class SendCertificatesBulkView(APIView):
    """
    Envía los certificados PDF almacenados en la BD (BLOB) por correo a los estudiantes.
    Usa threading para no bloquear el frontend.
    Al finalizar, envía un reporte al admin que inició la acción.
    """
    def post(self, request, *args, **kwargs):
        event_id = request.data.get('event_id')
        if not event_id:
            return Response({"error": "Falta event_id"}, status=status.HTTP_400_BAD_REQUEST)
            
        # Validar que existan certificados ANTES de lanzar el hilo
        certificates = GeneratedCertificate.objects.filter(evento_id=event_id)
        if not certificates.exists():
            return Response({"error": "No hay certificados generados para este evento. Genérelos primero."}, status=status.HTTP_400_BAD_REQUEST)

        # Email del admin que solicitó la acción (para el reporte)
        admin_email = request.user.email if request.user and request.user.email else None

        # Lanzar hilo en segundo plano
        import threading
        hilo = threading.Thread(
            target=_worker_send_certificates,
            args=(event_id, admin_email),
            daemon=True
        )
        hilo.start()

        return Response({
            "message": "El envío de certificados se está procesando en segundo plano. Recibirás un correo al finalizar.",
            "status": "background_processing"
        }, status=status.HTTP_200_OK)


def _worker_send_certificates(event_id, admin_email):
    """Worker que ejecuta el envío masivo de certificados en un hilo separado."""
    from django.core.mail import EmailMessage as DjangoEmailMessage
    from django.core.mail import send_mail as django_send_mail

    sent_count = 0
    error_count = 0
    errors = []
    evento_titulo = "Evento"

    try:
        certificates = GeneratedCertificate.objects.filter(evento_id=event_id).select_related('usuario', 'evento')

        for cert in certificates:
            try:
                user = cert.usuario
                evento_titulo = cert.evento.titulo
                if not user.email:
                    continue

                subject = f"🎓 Certificado de Asistencia: {cert.evento.titulo}"
                body = f"""
Hola {user.full_name},

Agradecemos sinceramente tu participación en el evento "{cert.evento.titulo}".

Adjunto a este correo encontrarás tu certificado oficial de asistencia en formato PDF.

Esperamos verte en nuestros próximos eventos.

Atentamente,
Universidad del Valle - Sistema SIGUE
"""
                email = DjangoEmailMessage(
                    subject=subject,
                    body=body.strip(),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    to=[user.email],
                )

                if cert.pdf_blob:
                    email.attach(cert.filename, cert.pdf_blob, 'application/pdf')
                    email.send(fail_silently=False)
                    sent_count += 1
                else:
                    raise Exception("El certificado no tiene contenido binario (PDF vacío).")

            except Exception as e:
                error_count += 1
                user_email_str = getattr(user, 'email', 'desconocido') if 'user' in dir() else 'desconocido'
                print(f"Error enviando certificado a {user_email_str}: {str(e)}")
                errors.append(f"{user_email_str}: {str(e)}")

    except Exception as e:
        print(f"Error crítico en hilo de certificados: {e}")
        errors.append(f"Error crítico: {str(e)}")

    # Enviar reporte al admin
    if admin_email:
        try:
            errores_str = "\n".join(errors[:10]) if errors else "Ninguno"
            django_send_mail(
                f'✅ Reporte de Envío de Certificados - {evento_titulo}',
                f'El envío masivo de certificados ha finalizado.\n\n'
                f'📊 Resumen:\n'
                f'  • Enviados exitosamente: {sent_count}\n'
                f'  • Fallidos: {error_count}\n\n'
                f'Errores:\n{errores_str}\n\n'
                f'— Sistema SIGUE',
                settings.DEFAULT_FROM_EMAIL,
                [admin_email],
                fail_silently=True
            )
        except Exception as e:
            print(f"Error enviando reporte al admin: {e}")

    print(f"[HILO CERTIFICADOS] Finalizado — Enviados: {sent_count}, Fallidos: {error_count}")


class CertificateViewSet(viewsets.ModelViewSet):
    queryset = GeneratedCertificate.objects.all().order_by('-created_at')
    from .serializers import GeneratedCertificateSerializer
    serializer_class = GeneratedCertificateSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        queryset = GeneratedCertificate.objects.all().order_by('-created_at')
        
        # Filter by Event
        event_id = self.request.query_params.get('event_id')
        if event_id:
            queryset = queryset.filter(evento_id=event_id)
        
        # If not admin, only see own certificates
        if user.role != 'Administrador':
            queryset = queryset.filter(usuario=user)
            
        return queryset

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        certificate = self.get_object()
        from django.http import HttpResponse
        response = HttpResponse(certificate.pdf_blob, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=\"{certificate.filename}\"'
        return response

class DownloadCertificatesZipView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ids = request.data.get('certificate_ids', []) # Standardized to certificate_ids per user request
        
        if not ids:
            return Response({"error": "No se seleccionaron certificados."}, status=status.HTTP_400_BAD_REQUEST)

        # Buscar certificados
        certificates = GeneratedCertificate.objects.filter(id__in=ids)
        
        if not certificates.exists():
            return Response({"error": "No se encontraron certificados válidos."}, status=status.HTTP_404_NOT_FOUND)

        # Crear el archivo ZIP en memoria
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for cert in certificates:
                if cert.pdf_blob:
                    filename = cert.filename or f"certificado_{cert.id}.pdf"
                    zip_file.writestr(filename, cert.pdf_blob)
        
        # Preparar respuesta
        zip_buffer.seek(0)
        from django.http import HttpResponse
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="certificados_seleccionados.zip"'
        return response
